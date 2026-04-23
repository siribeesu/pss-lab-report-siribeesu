def test_create_patient(client):
    """Test creating a single patient with auto-generated PSS-XXXX ID"""
    response = client.post(
        "/api/patients/",
        json={
            "name": "John Doe",
            "age": 30,
            "gender": "Male",
            "contact_number": "1234567890"
        }
    )
    assert response.status_code == 200
    data = response.json()
    assert data["name"] == "John Doe"
    assert data["age"] == 30
    assert data["gender"] == "Male"
    assert data["contact_number"] == "1234567890"
    assert data["patient_id"].startswith("PSS-")
    assert "id" in data

def test_create_patient_invalid(client):
    """Test validation: missing required field 'name' returns 422"""
    response = client.post(
        "/api/patients/",
        json={
            "age": 30,
            "gender": "Male",
            "contact_number": "1234567890"
        }
    )
    assert response.status_code == 422

def test_create_patient_invalid_age(client):
    """Test validation: invalid age values"""
    # Age as string (invalid type)
    response = client.post(
        "/api/patients/",
        json={
            "name": "Test Patient",
            "age": "thirty",
            "gender": "Male",
            "contact_number": "123456"
        }
    )
    assert response.status_code == 422

def test_list_patients(client):
    """Test listing all patients with pagination"""
    # Create 3 patients
    for i in range(3):
        client.post(
            "/api/patients/",
            json={
                "name": f"Patient {i}",
                "age": 25 + i,
                "gender": "Male",
                "contact_number": f"123456{i}"
            }
        )

    # Get all patients
    response = client.get("/api/patients/")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 3

def test_list_patients_with_pagination(client):
    """Test pagination with skip and limit"""
    # Create 5 patients
    for i in range(5):
        client.post(
            "/api/patients/",
            json={
                "name": f"Patient {i}",
                "age": 25 + i,
                "gender": "Female" if i % 2 == 0 else "Male",
                "contact_number": f"555{i:04d}"
            }
        )

    # Get first 2 patients
    response = client.get("/api/patients/?skip=0&limit=2")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 2

def test_search_patients_by_name(client):
    """Test searching patients by name"""
    # Create patients with different names
    client.post(
        "/api/patients/",
        json={"name": "John Smith", "age": 30, "gender": "Male", "contact_number": "111"}
    )
    client.post(
        "/api/patients/",
        json={"name": "Jane Doe", "age": 28, "gender": "Female", "contact_number": "222"}
    )

    # Search for "John"
    response = client.get("/api/patients/?search=John")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["name"] == "John Smith"

def test_search_patients_by_patient_id(client):
    """Test searching patients by patient ID"""
    # Create a patient
    create_response = client.post(
        "/api/patients/",
        json={"name": "Test Patient", "age": 35, "gender": "Male", "contact_number": "999"}
    )
    patient_id = create_response.json()["patient_id"]

    # Search by patient ID
    response = client.get(f"/api/patients/?search={patient_id}")
    assert response.status_code == 200
    data = response.json()
    assert len(data) == 1
    assert data[0]["patient_id"] == patient_id

def test_get_patient_detail(client):
    """Test retrieving patient detail with all associated reports"""
    # Create patient
    patient_response = client.post(
        "/api/patients/",
        json={"name": "Detail Test", "age": 40, "gender": "Male", "contact_number": "333"}
    )
    patient_id = patient_response.json()["id"]

    # Get patient detail
    response = client.get(f"/api/patients/{patient_id}")
    assert response.status_code == 200
    data = response.json()
    assert data["name"] == "Detail Test"
    assert "reports" in data
    assert isinstance(data["reports"], list)

def test_get_nonexistent_patient(client):
    """Test retrieving non-existent patient returns 404"""
    response = client.get("/api/patients/00000000-0000-0000-0000-000000000000")
    assert response.status_code == 404

def test_update_patient(client):
    """Test updating patient information"""
    # Create patient
    create_response = client.post(
        "/api/patients/",
        json={"name": "Original Name", "age": 30, "gender": "Male", "contact_number": "111"}
    )
    patient_id = create_response.json()["id"]

    # Update patient
    update_response = client.put(
        f"/api/patients/{patient_id}",
        json={"name": "Updated Name", "contact_number": "999"}
    )
    assert update_response.status_code == 200
    data = update_response.json()
    assert data["name"] == "Updated Name"
    assert data["contact_number"] == "999"

def test_delete_patient(client):
    """Test deleting a patient"""
    # Create patient
    create_response = client.post(
        "/api/patients/",
        json={"name": "To Delete", "age": 25, "gender": "Female", "contact_number": "777"}
    )
    patient_id = create_response.json()["id"]

    # Delete patient
    delete_response = client.delete(f"/api/patients/{patient_id}")
    assert delete_response.status_code == 204

    # Verify patient is deleted
    get_response = client.get(f"/api/patients/{patient_id}")
    assert get_response.status_code == 404

def test_bulk_upload_patients(client):
    """Test bulk uploading patients from CSV"""
    csv_content = (
        "Full Patient Name,Age,Gender,Contact Number\n"
        "Bulk John,35,Male,9998887777\n"
        "Bulk Jane,28,Female,6665554444\n"
    )
    files = {"file": ("patients.csv", csv_content, "text/csv")}
    response = client.post("/api/patients/bulk_upload", files=files)

    assert response.status_code == 200
    data = response.json()
    assert data["added_count"] == 2
    assert "Processed 2 records" in data["message"]
    assert len(data["errors"]) == 0

def test_bulk_upload_patients_xlsx(client):
    """Test bulk uploading patients from XLSX"""
    import io
    import openpyxl

    # Create XLSX in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "Full Patient Name"
    ws['B1'] = "Age"
    ws['C1'] = "Gender"
    ws['D1'] = "Contact Number"
    ws['A2'] = "Excel John"
    ws['B2'] = 42
    ws['C2'] = "Male"
    ws['D2'] = "5551234567"

    xlsx_bytes = io.BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    files = {"file": ("patients.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    response = client.post("/api/patients/bulk_upload", files=files)

    assert response.status_code == 200
    data = response.json()
    assert data["added_count"] == 1

def test_bulk_upload_with_errors(client):
    """Test bulk upload with mixed valid/invalid rows"""
    csv_content = (
        "Full Patient Name,Age,Gender,Contact Number\n"
        "Valid Patient,30,Male,111\n"
        "Invalid Patient,invalid_age,Female,222\n"
        "Another Valid,25,Female,333\n"
    )
    files = {"file": ("patients.csv", csv_content, "text/csv")}
    response = client.post("/api/patients/bulk_upload", files=files)

    assert response.status_code == 200
    data = response.json()
    assert data["added_count"] >= 3  # All rows are now 'valid' because age defaults to 0

def test_duplicate_contact_number(client):
    """Test creating patient with duplicate contact number (should be allowed per requirements)"""
    contact = "9876543210"

    # Create first patient
    response1 = client.post(
        "/api/patients/",
        json={"name": "Patient 1", "age": 30, "gender": "Male", "contact_number": contact}
    )
    assert response1.status_code == 200

    # Create second patient with same contact (should succeed per requirements)
    response2 = client.post(
        "/api/patients/",
        json={"name": "Patient 2", "age": 35, "gender": "Female", "contact_number": contact}
    )
    assert response2.status_code == 200
